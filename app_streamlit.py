"""
MLI Processing Tool
Mean Linear Intercept (MLI) and Mean Wall Thickness (MWT) Analysis
"""

import os
import numpy as np
import cv2
import pandas as pd
import math
import streamlit as st
from pathlib import Path
import tempfile
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from PIL import Image
import io
import zipfile


# Set page config - must be the first Streamlit command
st.set_page_config(
    page_title="MLI Processing Tool",
    page_icon="🔬",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for better styling
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: 700;
        color: #1f77b4;
        margin-bottom: 0.5rem;
    }
    .sub-header {
        font-size: 1.2rem;
        color: #666;
        margin-bottom: 2rem;
    }
    .metric-card {
        background-color: #f0f2f6;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 0.5rem 0;
    }
    .success-box {
        background-color: #d4edda;
        border-left: 4px solid #28a745;
        padding: 1rem;
        border-radius: 0.25rem;
        margin: 1rem 0;
    }
    .info-box {
        background-color: #d1ecf1;
        border-left: 4px solid #17a2b8;
        padding: 1rem;
        border-radius: 0.25rem;
        margin: 1rem 0;
    }
    .stButton>button {
        width: 100%;
        background-color: #1f77b4;
        color: white;
        font-weight: 600;
        border-radius: 0.5rem;
        padding: 0.5rem 1rem;
        border: none;
        transition: all 0.3s;
    }
    .stButton>button:hover {
        background-color: #155a8a;
        transform: translateY(-2px);
        box-shadow: 0 4px 8px rgba(0,0,0,0.2);
    }
</style>
""", unsafe_allow_html=True)


class HuangThresholding:
    """Huang's fuzzy thresholding method implementation"""

    def __init__(self, data):
        self.data = data
        self.first_bin, self.last_bin = self.find_bin_limits()
        self.term = 1.0 / max(1, self.last_bin - self.first_bin)
        self.mu_0, self.mu_1 = self.calculate_mu()

    def find_bin_limits(self):
        non_zero_indices = np.nonzero(self.data)[0]
        first_bin = non_zero_indices[0]
        last_bin = non_zero_indices[-1]
        return first_bin, last_bin

    def calculate_mu(self):
        indices = np.arange(len(self.data))
        num_pix_cumsum = np.cumsum(self.data)
        sum_pix_cumsum = np.cumsum(indices * self.data)
        mu_0 = sum_pix_cumsum / \
            np.where(num_pix_cumsum == 0, 1, num_pix_cumsum)

        num_pix_cumsum_rev = np.cumsum(self.data[::-1])[::-1]
        sum_pix_cumsum_rev = np.cumsum((indices[::-1]) * self.data[::-1])[::-1]
        mu_1 = sum_pix_cumsum_rev / \
            np.where(num_pix_cumsum_rev == 0, 1, num_pix_cumsum_rev)

        return mu_0, mu_1

    def calculate_entropy(self, it):
        ent = 0.0
        for ih in range(it):
            mu_x = 1.0 / (1.0 + self.term * abs(ih - self.mu_0[it]))
            if not (mu_x < 1e-6 or mu_x > 1 - 1e-6):
                ent -= self.data[ih] * (
                    mu_x * math.log(mu_x) + (1.0 - mu_x) * math.log(1.0 - mu_x)
                )

        for ih in range(it + 1, len(self.data)):
            mu_x = 1.0 / (1.0 + self.term * abs(ih - self.mu_1[it]))
            if not (mu_x < 1e-6 or mu_x > 1 - 1e-6):
                ent -= self.data[ih] * (
                    mu_x * math.log(mu_x) + (1.0 - mu_x) * math.log(1.0 - mu_x)
                )

        return ent

    def find_threshold(self):
        threshold = -1
        min_ent = float("inf")
        for it in range(self.first_bin, self.last_bin + 1):
            ent = self.calculate_entropy(it)
            if ent < min_ent:
                min_ent = ent
                threshold = it
        return threshold


def create_test_lines(image_width, image_height, line_spacing, orientation='horizontal'):
    """Create test grid lines"""
    image = np.ones((image_height, image_width, 3), np.uint8) * 255
    if orientation == 'horizontal':
        for y in range(0, image_height, line_spacing):
            cv2.line(image, (0, y), (image_width, y), (0, 0, 255), 1)
    elif orientation == 'vertical':
        for x in range(0, image_width, line_spacing):
            cv2.line(image, (x, 0), (x, image_height), (0, 0, 255), 1)
    return image


def contrast_threshold(image_path):
    """Apply contrast-enhanced thresholding"""
    image = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
    hist, _ = np.histogram(image, bins=np.arange(257), density=False)
    huang = HuangThresholding(hist)
    threshold = huang.find_threshold()
    _, thresholded_image = cv2.threshold(
        image, threshold, 255, cv2.THRESH_BINARY)
    return thresholded_image


def overlay_images(base_image, overlay_image, measure='mli'):
    """Overlay grid lines on base image"""
    if len(base_image.shape) == 2:
        base_image = cv2.cvtColor(base_image, cv2.COLOR_GRAY2BGR)

    if measure == 'mli':
        base_image_processed = cv2.bitwise_not(base_image)
    elif measure == 'mwt':
        base_image_processed = base_image.copy()
    else:
        raise ValueError("Invalid measure type. Must be 'mli' or 'mwt'.")

    lower_red = np.array([0, 0, 150])
    upper_red = np.array([100, 100, 255])
    mask = cv2.inRange(overlay_image, lower_red, upper_red)
    combined_image = overlay_image.copy()
    black_color = np.array([0, 0, 0], dtype=np.uint8)
    combined_image[mask > 0] = black_color
    final_image = cv2.addWeighted(
        base_image_processed, 1, combined_image, 1, 0)
    return final_image


def measure_chords(image, pixel_width, pixel_height):
    """Measure chord lengths"""
    if len(image.shape) == 3:
        image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    inverted_image = cv2.bitwise_not(image)
    image_height, image_width = inverted_image.shape
    contours, _ = cv2.findContours(
        inverted_image, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    mask = np.zeros_like(inverted_image)
    chord_lengths = []

    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        if x == 0 or y == 0 or (x + w) == image_width or (y + h) == image_height:
            continue
        cv2.drawContours(mask, [contour], -1, 255, thickness=cv2.FILLED)

        if w > h:
            length = w
            orientation = 'Horizontal'
        elif h > w:
            length = h
            orientation = 'Vertical'
        else:
            continue

        chord_info = {
            'x': x,
            'y': y,
            'Length': length * pixel_width if orientation == 'Horizontal' else length * pixel_height,
            'Orientation': orientation
        }
        chord_lengths.append(chord_info)

    final_image = cv2.bitwise_and(inverted_image, mask)
    final_image = cv2.bitwise_not(final_image)
    return chord_lengths, final_image


def convert_black_to_red(image):
    """Convert black lines to red for visualization"""
    if len(image.shape) == 2:
        image = cv2.cvtColor(image, cv2.COLOR_GRAY2BGR)

    black_color = np.array([0, 0, 0], dtype=np.uint8)
    red_color = np.array([0, 0, 255], dtype=np.uint8)
    image[np.all(image == black_color, axis=-1)] = red_color
    return image


def convert_gray_to_bgr(image):
    """Convert grayscale to BGR"""
    if len(image.shape) == 2:
        image = cv2.cvtColor(image, cv2.COLOR_GRAY2BGR)
    return image


def create_readme_file(output_image_dir):
    """Create a README file documenting the folder structure"""
    readme_content = """# Output Folder Structure

This directory contains the analysis results for a single image, organized into the following folders:

## 01_preprocessing/
- **contrast_threshold.tif**: Binary thresholded image (black=tissue, white=airspace)
  - Used as the foundation for all subsequent measurements

## 02_contours/
- **all_contours_[filename].tif**: All detected contours in the image
- **final_contours_[filename].tif**: Filtered contours (excluding edge-touching)
- **final_contours_filled_[filename].tif**: Same as above but with filled contours (for visualization)
  - Red areas indicate the alveolar spaces that were measured

## 03_mli/ (Mean Linear Intercept Analysis)
### grids/
- **horizontal_grid_mli.tif**: Horizontal test lines overlaid on inverted image
- **vertical_grid_mli.tif**: Vertical test lines overlaid on inverted image
  - Black segments represent airspace intercepts that were measured

### overlays/
- **overlay_horizontal_mli.tif**: Horizontal grid overlay on thresholded image
- **overlay_vertical_mli.tif**: Vertical grid overlay on thresholded image
  - Red lines show the measured intercepts overlaid on the original tissue

## 04_mwt/ (Mean Wall Thickness Analysis)
### grids/
- **horizontal_grid_mwt.tif**: Horizontal test lines showing tissue wall intercepts
- **vertical_grid_mwt.tif**: Vertical test lines showing tissue wall intercepts
  - Black segments represent wall thickness measurements

### overlays/
- **overlay_horizontal_mwt.tif**: Horizontal grid overlay for wall thickness
- **overlay_vertical_mwt.tif**: Vertical grid overlay for wall thickness
  - Red lines show measured wall segments on the tissue

## 05_data/
- **chord_lengths_mli.csv**: Individual MLI measurements
  - Columns: x, y, Length, Orientation, Label, Measure_Type
- **chord_lengths_mwt.csv**: Individual MWT measurements
  - Columns: x, y, Length, Orientation, Label, Measure_Type

---

## Workflow
1. Image is thresholded to separate tissue from airspace (01_preprocessing)
2. Contours are detected and filtered (02_contours)
3. Grid lines are overlaid and intercepts measured (03_mli and 04_mwt)
4. Raw measurements are saved to CSV files (05_data)
5. Summary statistics are calculated at the slide set level (parent directory)

## Quality Control
- Review thresholded image to ensure proper tissue/airspace separation
- Check filled contours to verify which airspaces were included in measurements
- Examine overlay images to confirm grid lines are properly aligned with tissue structures
"""

    readme_path = os.path.join(output_image_dir, 'README.md')
    with open(readme_path, 'w') as f:
        f.write(readme_content)


def measure_areas_of_spaces(thresholded_image, output_image_dir, base_name):
    """Measure areas of spaces in the image"""
    if len(thresholded_image.shape) == 3:
        thresholded_image = cv2.cvtColor(thresholded_image, cv2.COLOR_BGR2GRAY)

    contours, _ = cv2.findContours(
        thresholded_image, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    thresholded_image_copy_all = cv2.cvtColor(
        thresholded_image, cv2.COLOR_GRAY2BGR)
    cv2.drawContours(thresholded_image_copy_all, contours, -1, (0, 0, 255), 1)

    all_contours_path = os.path.join(
        output_image_dir, f'all_contours_{base_name}.tif')
    cv2.imwrite(all_contours_path, thresholded_image_copy_all)

    areas = []
    height, width = thresholded_image.shape
    passed_contours = []

    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        if x > 0 and y > 0 and (x + w) < width and (y + h) < height:
            area = cv2.contourArea(contour)
            areas.append(area)
            passed_contours.append(contour)

    thresholded_image_copy_final = cv2.cvtColor(
        thresholded_image, cv2.COLOR_GRAY2BGR)
    cv2.drawContours(thresholded_image_copy_final,
                     passed_contours, -1, (0, 0, 255), 1)

    final_contours_path = os.path.join(
        output_image_dir, f'final_contours_{base_name}.tif')
    cv2.imwrite(final_contours_path, thresholded_image_copy_final)

    # Create filled version for preview
    thresholded_image_copy_filled = cv2.cvtColor(
        thresholded_image, cv2.COLOR_GRAY2BGR)
    cv2.drawContours(thresholded_image_copy_filled,
                     passed_contours, -1, (0, 0, 255), thickness=cv2.FILLED)

    final_contours_filled_path = os.path.join(
        output_image_dir, f'final_contours_filled_{base_name}.tif')
    cv2.imwrite(final_contours_filled_path, thresholded_image_copy_filled)

    return areas, thresholded_image_copy_filled


def process_single_image(image_path, output_dir, input_dir_name, line_spacing,
                         pixel_width, pixel_height, progress_callback=None):
    """Process a single image and return results"""
    filename = os.path.basename(image_path)
    base_name = os.path.splitext(filename)[0]
    output_subdir = os.path.join(output_dir, input_dir_name)
    output_image_dir = os.path.join(output_subdir, base_name)

    # Create organized folder structure
    preprocessing_dir = os.path.join(output_image_dir, '01_preprocessing')
    contours_dir = os.path.join(output_image_dir, '02_contours')
    mli_dir = os.path.join(output_image_dir, '03_mli')
    mli_grids_dir = os.path.join(mli_dir, 'grids')
    mli_overlays_dir = os.path.join(mli_dir, 'overlays')
    mwt_dir = os.path.join(output_image_dir, '04_mwt')
    mwt_grids_dir = os.path.join(mwt_dir, 'grids')
    mwt_overlays_dir = os.path.join(mwt_dir, 'overlays')
    data_dir = os.path.join(output_image_dir, '05_data')

    for directory in [preprocessing_dir, contours_dir, mli_grids_dir, mli_overlays_dir,
                      mwt_grids_dir, mwt_overlays_dir, data_dir]:
        if not os.path.exists(directory):
            os.makedirs(directory)

    # Load the image
    image = cv2.imread(image_path)
    if image is None:
        return None, f"Failed to load image: {filename}"

    image_width, image_height = image.shape[1], image.shape[0]

    # Create test lines
    horizontal_lines = create_test_lines(
        image_width, image_height, line_spacing, orientation='horizontal')
    vertical_lines = create_test_lines(
        image_width, image_height, line_spacing, orientation='vertical')

    # Apply thresholding
    thresholded_image = contrast_threshold(image_path)
    cv2.imwrite(os.path.join(preprocessing_dir,
                'contrast_threshold.tif'), thresholded_image)

    if thresholded_image is None:
        return None, f"Thresholding failed for: {filename}"

    # Measure areas
    areas, final_contours_image = measure_areas_of_spaces(
        thresholded_image, contours_dir, base_name)
    total_area_pixels = sum(areas)
    total_area_microns = total_area_pixels * pixel_width * pixel_height

    label = os.path.join(input_dir_name, base_name)
    image_summary = {
        'Label': label,
        'Total_Area_Pixels': total_area_pixels,
        'Total_Area_Microns': total_area_microns
    }

    thresholded_image_bgr = convert_gray_to_bgr(thresholded_image)

    # Process MLI
    combined_image_horizontal_mli = overlay_images(
        thresholded_image, horizontal_lines, measure='mli')
    combined_image_vertical_mli = overlay_images(
        thresholded_image, vertical_lines, measure='mli')

    chord_lengths_horizontal_mli, final_image_horizontal_mli = measure_chords(
        combined_image_horizontal_mli, pixel_width, pixel_height)
    chord_lengths_vertical_mli, final_image_vertical_mli = measure_chords(
        combined_image_vertical_mli, pixel_width, pixel_height)

    cv2.imwrite(os.path.join(mli_grids_dir,
                'horizontal_grid_mli.tif'), final_image_horizontal_mli)
    cv2.imwrite(os.path.join(mli_grids_dir,
                'vertical_grid_mli.tif'), final_image_vertical_mli)

    final_image_horizontal_mli_red = convert_black_to_red(
        final_image_horizontal_mli)
    final_image_vertical_mli_red = convert_black_to_red(
        final_image_vertical_mli)

    overlay_horizontal_mli = cv2.addWeighted(
        thresholded_image_bgr, 0.5, final_image_horizontal_mli_red, 0.5, 0)
    overlay_vertical_mli = cv2.addWeighted(
        thresholded_image_bgr, 0.5, final_image_vertical_mli_red, 0.5, 0)

    cv2.imwrite(os.path.join(mli_overlays_dir,
                'overlay_horizontal_mli.tif'), overlay_horizontal_mli)
    cv2.imwrite(os.path.join(mli_overlays_dir,
                'overlay_vertical_mli.tif'), overlay_vertical_mli)

    # Save MLI data
    df_mli = pd.DataFrame(chord_lengths_horizontal_mli +
                          chord_lengths_vertical_mli)
    if not df_mli.empty:
        df_mli = df_mli.sort_values(by=['Orientation'])
        df_mli['Length'] = df_mli['Length'].round(2)
        df_mli['x'] = pd.to_numeric(df_mli['x'], downcast='integer')
        df_mli['y'] = pd.to_numeric(df_mli['y'], downcast='integer')
        df_mli['Length'] = pd.to_numeric(df_mli['Length'])
        df_mli['Label'] = label
        df_mli['Measure_Type'] = 'MLI'
        df_mli.to_csv(os.path.join(data_dir,
                      'chord_lengths_mli.csv'), index=False)

    # Process MWT
    combined_image_horizontal_mwt = overlay_images(
        thresholded_image, horizontal_lines, measure='mwt')
    combined_image_vertical_mwt = overlay_images(
        thresholded_image, vertical_lines, measure='mwt')

    chord_lengths_horizontal_mwt, final_image_horizontal_mwt = measure_chords(
        combined_image_horizontal_mwt, pixel_width, pixel_height)
    chord_lengths_vertical_mwt, final_image_vertical_mwt = measure_chords(
        combined_image_vertical_mwt, pixel_width, pixel_height)

    cv2.imwrite(os.path.join(mwt_grids_dir,
                'horizontal_grid_mwt.tif'), final_image_horizontal_mwt)
    cv2.imwrite(os.path.join(mwt_grids_dir,
                'vertical_grid_mwt.tif'), final_image_vertical_mwt)

    final_image_horizontal_mwt_red = convert_black_to_red(
        final_image_horizontal_mwt)
    final_image_vertical_mwt_red = convert_black_to_red(
        final_image_vertical_mwt)

    overlay_horizontal_mwt = cv2.addWeighted(
        thresholded_image_bgr, 0.5, final_image_horizontal_mwt_red, 0.5, 0)
    overlay_vertical_mwt = cv2.addWeighted(
        thresholded_image_bgr, 0.5, final_image_vertical_mwt_red, 0.5, 0)

    cv2.imwrite(os.path.join(mwt_overlays_dir,
                'overlay_horizontal_mwt.tif'), overlay_horizontal_mwt)
    cv2.imwrite(os.path.join(mwt_overlays_dir,
                'overlay_vertical_mwt.tif'), overlay_vertical_mwt)

    # Save MWT data
    df_mwt = pd.DataFrame(chord_lengths_horizontal_mwt +
                          chord_lengths_vertical_mwt)
    if not df_mwt.empty:
        df_mwt = df_mwt.sort_values(by=['Orientation'])
        df_mwt['Length'] = df_mwt['Length'].round(2)
        df_mwt['x'] = pd.to_numeric(df_mwt['x'], downcast='integer')
        df_mwt['y'] = pd.to_numeric(df_mwt['y'], downcast='integer')
        df_mwt['Length'] = pd.to_numeric(df_mwt['Length'])
        df_mwt['Label'] = label
        df_mwt['Measure_Type'] = 'MWT'
        df_mwt.to_csv(os.path.join(data_dir,
                      'chord_lengths_mwt.csv'), index=False)

    # Create README file documenting the folder structure
    create_readme_file(output_image_dir)

    return {
        'image_summary': image_summary,
        'output_dir': output_image_dir,
        'filename': filename,
        'original': image,
        'overlay_mli': overlay_horizontal_mli,
        'overlay_mli_vertical': overlay_vertical_mli,
        'overlay_mwt': overlay_horizontal_mwt,
        'overlay_mwt_vertical': overlay_vertical_mwt,
        'thresholded': thresholded_image_bgr,
        'final_contours': final_contours_image
    }, None


def process_csv_files(output_directory, slide_ROI_separator):
    """Process and aggregate CSV results"""
    csv_files = []
    for root, dirs, files in os.walk(output_directory):
        for file in files:
            if file in ["chord_lengths_mli.csv", "chord_lengths_mwt.csv"]:
                csv_files.append(os.path.join(root, file))

    if not csv_files:
        return None

    df_list = []
    for file in csv_files:
        df = pd.read_csv(file)
        if not df.empty:
            df_list.append(df)

    if not df_list:
        return None

    slide_set = pd.concat(df_list, ignore_index=True)
    slide_set[['input_dir_name', 'base_name']
              ] = slide_set['Label'].str.split(os.sep, n=1, expand=True)
    slide_set[['slide_num', 'ROI']] = slide_set['base_name'].str.split(
        slide_ROI_separator, expand=True)
    slide_set = slide_set.astype({'slide_num': 'str', 'ROI': 'str'})

    slide_set_mli = slide_set[slide_set['Measure_Type'] == 'MLI']
    slide_set_mwt = slide_set[slide_set['Measure_Type'] == 'MWT']

    mli_vert_grid_avg = (slide_set_mli[slide_set_mli['Orientation'] == "Vertical"]
                         .groupby(['input_dir_name', 'slide_num', 'ROI'])
                         .agg(vert_grid_avg_mli=('Length', 'mean'))
                         .reset_index())

    mli_horiz_grid_avg = (slide_set_mli[slide_set_mli['Orientation'] == "Horizontal"]
                          .groupby(['input_dir_name', 'slide_num', 'ROI'])
                          .agg(horiz_grid_avg_mli=('Length', 'mean'))
                          .reset_index())

    mwt_vert_grid_avg = (slide_set_mwt[slide_set_mwt['Orientation'] == "Vertical"]
                         .groupby(['input_dir_name', 'slide_num', 'ROI'])
                         .agg(vert_grid_avg_mwt=('Length', 'mean'))
                         .reset_index())

    mwt_horiz_grid_avg = (slide_set_mwt[slide_set_mwt['Orientation'] == "Horizontal"]
                          .groupby(['input_dir_name', 'slide_num', 'ROI'])
                          .agg(horiz_grid_avg_mwt=('Length', 'mean'))
                          .reset_index())

    slide_set_grid_avg_mli = pd.merge(mli_horiz_grid_avg, mli_vert_grid_avg,
                                      on=['input_dir_name', 'slide_num', 'ROI'])
    slide_set_grid_avg_mwt = pd.merge(mwt_horiz_grid_avg, mwt_vert_grid_avg,
                                      on=['input_dir_name', 'slide_num', 'ROI'])

    # Load image summaries
    image_summaries_files = []
    for root, dirs, files in os.walk(output_directory):
        for file in files:
            if file == "image_summaries.csv":
                image_summaries_files.append(os.path.join(root, file))

    if image_summaries_files:
        image_summaries_list = [pd.read_csv(file)
                                for file in image_summaries_files]
        df_image_summaries = pd.concat(image_summaries_list, ignore_index=True)
        df_image_summaries[['input_dir_name', 'base_name']
                           ] = df_image_summaries['Label'].str.split(os.sep, n=1, expand=True)
        df_image_summaries[['slide_num', 'ROI']] = df_image_summaries['base_name'].str.split(
            slide_ROI_separator, expand=True)
        df_image_summaries = df_image_summaries.astype(
            {'slide_num': 'str', 'ROI': 'str'})

        slide_set_grid_avg_mli = pd.merge(slide_set_grid_avg_mli, df_image_summaries,
                                          on=['input_dir_name', 'slide_num', 'ROI'])
        slide_set_grid_avg_mwt = pd.merge(slide_set_grid_avg_mwt, df_image_summaries,
                                          on=['input_dir_name', 'slide_num', 'ROI'])

    slide_set_grid_avg_mli['mli'] = slide_set_grid_avg_mli[[
        'horiz_grid_avg_mli', 'vert_grid_avg_mli']].mean(axis=1)
    slide_set_grid_avg_mwt['mwt'] = slide_set_grid_avg_mwt[[
        'horiz_grid_avg_mwt', 'vert_grid_avg_mwt']].mean(axis=1)

    slide_set_grid_avg = pd.merge(
        slide_set_grid_avg_mli,
        slide_set_grid_avg_mwt[['input_dir_name', 'slide_num', 'ROI',
                                'horiz_grid_avg_mwt', 'vert_grid_avg_mwt', 'mwt']],
        on=['input_dir_name', 'slide_num', 'ROI'],
        how='outer'
    )

    slide_set_grid_avg.to_csv(os.path.join(
        output_directory, 'slide_set_grid_avg.csv'), index=False)

    avg_per_slide = (slide_set_grid_avg.groupby(['input_dir_name', 'slide_num'])
                     .agg(mli_avg_per_slide=('mli', 'mean'),
                          mwt_avg_per_slide=('mwt', 'mean'),
                          avg_total_area_microns=('Total_Area_Microns', 'mean'))
                     .reset_index())

    return avg_per_slide


def write_to_excel(df, file_path):
    """Write DataFrame to formatted Excel file"""
    wb = Workbook()
    ws = wb.active

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD",
                              end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(file_path)


def main():
    """Main Streamlit application"""

    # Header
    st.markdown('<div class="main-header">🔬 MLI Processing Tool</div>',
                unsafe_allow_html=True)
    st.markdown('<div class="sub-header">Mean Linear Intercept (MLI) and Mean Wall Thickness (MWT) Analysis</div>',
                unsafe_allow_html=True)

    # Sidebar for parameters
    st.sidebar.header("⚙️ Analysis Parameters")

    # Parameter inputs
    line_spacing = st.sidebar.number_input(
        "Line Spacing (pixels)",
        min_value=1,
        max_value=100,
        value=15,
        help="Distance between grid lines in pixels"
    )

    pixel_width = st.sidebar.number_input(
        "Pixel Width (µm)",
        min_value=0.01,
        max_value=10.0,
        value=0.57,
        format="%.4f",
        help="Width of a single pixel in micrometers"
    )

    pixel_height = st.sidebar.number_input(
        "Pixel Height (µm)",
        min_value=0.01,
        max_value=10.0,
        value=0.57,
        format="%.4f",
        help="Height of a single pixel in micrometers"
    )

    slide_roi_separator = st.sidebar.text_input(
        "Slide-ROI Separator",
        value="_",
        help="Character that separates slide number from ROI in filename"
    )

    # Calculate and display spacing
    horizontal_spacing = line_spacing * pixel_width
    vertical_spacing = line_spacing * pixel_height

    st.sidebar.markdown("---")
    st.sidebar.markdown("### 📏 Calculated Spacing")
    st.sidebar.metric("Horizontal", f"{horizontal_spacing:.2f} µm")
    st.sidebar.metric("Vertical", f"{vertical_spacing:.2f} µm")

    # Main content area
    tab1, tab2, tab3 = st.tabs(["📁 Upload & Process", "📊 Results", "ℹ️ About"])

    with tab1:
        st.header("Upload Images for Analysis")

        col1, col2 = st.columns([2, 1])

        with col1:
            uploaded_files = st.file_uploader(
                "Select image files (TIF, TIFF, PNG, JPG, JPEG)",
                type=['tif', 'tiff', 'png', 'jpg', 'jpeg'],
                accept_multiple_files=True,
                help="You can select multiple image files at once"
            )

        with col2:
            output_dir_name = st.text_input(
                "Output Directory Name",
                value="mli_output",
                help="Name for the output directory"
            )

        if uploaded_files:
            st.success(
                f"✅ {len(uploaded_files)} file(s) uploaded successfully")

            # Show preview of uploaded files
            with st.expander("📋 View Uploaded Files"):
                for idx, file in enumerate(uploaded_files, 1):
                    st.text(f"{idx}. {file.name} ({file.size / 1024:.1f} KB)")

            if st.button("🚀 Start Processing", type="primary"):
                # Create temporary directory for processing
                with tempfile.TemporaryDirectory() as temp_dir:
                    input_temp_dir = os.path.join(temp_dir, "input")
                    output_temp_dir = os.path.join(temp_dir, "output")
                    os.makedirs(input_temp_dir)
                    os.makedirs(output_temp_dir)

                    # Save uploaded files
                    for uploaded_file in uploaded_files:
                        file_path = os.path.join(
                            input_temp_dir, uploaded_file.name)
                        with open(file_path, "wb") as f:
                            f.write(uploaded_file.getbuffer())

                    # Process images
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    results_container = st.container()

                    image_summaries = []
                    processed_count = 0
                    total_files = len(uploaded_files)

                    for idx, uploaded_file in enumerate(uploaded_files):
                        status_text.text(f"Processing: {uploaded_file.name}")

                        file_path = os.path.join(
                            input_temp_dir, uploaded_file.name)
                        result, error = process_single_image(
                            file_path,
                            output_temp_dir,
                            output_dir_name,
                            line_spacing,
                            pixel_width,
                            pixel_height
                        )

                        if result:
                            image_summaries.append(result['image_summary'])
                            processed_count += 1

                            # Show preview of processed image
                            with results_container.expander(f"✅ {uploaded_file.name}"):
                                # First row: Original, Thresholded, Final Contours
                                col1, col2, col3 = st.columns(3)
                                with col1:
                                    st.image(
                                        result['original'], caption="Original", use_container_width=True)
                                with col2:
                                    st.image(
                                        result['thresholded'], caption="Thresholded", use_container_width=True)
                                with col3:
                                    st.image(
                                        result['final_contours'], caption="Final Air Space", use_container_width=True)

                                # Second row: MLI overlays
                                st.markdown("**MLI Analysis:**")
                                col4, col5 = st.columns(2)
                                with col4:
                                    st.image(
                                        result['overlay_mli'], caption="MLI Horizontal Grid", use_container_width=True)
                                with col5:
                                    st.image(
                                        result['overlay_mli_vertical'], caption="MLI Vertical Grid", use_container_width=True)

                                # Third row: MWT overlays
                                st.markdown("**MWT Analysis:**")
                                col6, col7 = st.columns(2)
                                with col6:
                                    st.image(
                                        result['overlay_mwt'], caption="MWT Horizontal Grid", use_container_width=True)
                                with col7:
                                    st.image(
                                        result['overlay_mwt_vertical'], caption="MWT Vertical Grid", use_container_width=True)
                        else:
                            st.error(f"❌ {error}")

                        progress_bar.progress((idx + 1) / total_files)

                    status_text.text("Processing complete!")

                    # Save image summaries
                    if image_summaries:
                        output_subdir = os.path.join(
                            output_temp_dir, output_dir_name)
                        df_image_summaries = pd.DataFrame(image_summaries)
                        image_summaries_csv_path = os.path.join(
                            output_subdir, 'image_summaries.csv')
                        df_image_summaries.to_csv(
                            image_summaries_csv_path, index=False)

                        # Process CSV files
                        avg_per_slide = process_csv_files(
                            output_temp_dir, slide_roi_separator)

                        if avg_per_slide is not None:
                            excel_path = os.path.join(
                                output_temp_dir, 'avg_per_slide.xlsx')
                            write_to_excel(avg_per_slide, excel_path)

                            # Store results in session state
                            st.session_state['results'] = avg_per_slide
                            st.session_state['output_dir'] = output_temp_dir
                            st.session_state['image_summaries'] = df_image_summaries

                            st.balloons()
                            st.success(
                                f"🎉 Successfully processed {processed_count} out of {total_files} images!")

                            # Create download zip
                            zip_buffer = io.BytesIO()
                            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                                for root, dirs, files in os.walk(output_temp_dir):
                                    for file in files:
                                        file_path = os.path.join(root, file)
                                        arcname = os.path.relpath(
                                            file_path, output_temp_dir)
                                        zip_file.write(file_path, arcname)

                            zip_buffer.seek(0)
                            st.download_button(
                                label="📥 Download All Results (ZIP)",
                                data=zip_buffer,
                                file_name=f"{output_dir_name}_results.zip",
                                mime="application/zip"
                            )

    with tab2:
        st.header("Analysis Results")

        if 'results' in st.session_state and st.session_state['results'] is not None:
            results_df = st.session_state['results']

            # Display metrics
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total Slides", len(results_df))
            with col2:
                avg_mli = results_df['mli_avg_per_slide'].mean()
                st.metric("Average MLI", f"{avg_mli:.2f} µm")
            with col3:
                avg_mwt = results_df['mwt_avg_per_slide'].mean()
                st.metric("Average MWT", f"{avg_mwt:.2f} µm")

            st.markdown("---")

            # Display results table
            st.subheader("📋 Results Summary")
            st.dataframe(results_df, use_container_width=True)

            # Download individual files
            col1, col2 = st.columns(2)
            with col1:
                csv_buffer = results_df.to_csv(index=False).encode()
                st.download_button(
                    label="📥 Download Results (CSV)",
                    data=csv_buffer,
                    file_name="avg_per_slide.csv",
                    mime="text/csv"
                )

            # Display image summaries if available
            if 'image_summaries' in st.session_state:
                st.subheader("📸 Image Summaries")
                st.dataframe(
                    st.session_state['image_summaries'], use_container_width=True)
        else:
            st.info(
                "📊 No results yet. Process some images in the 'Upload & Process' tab first.")

    with tab3:
        st.header("About This Tool")

        # Section 1: What This App Does and How to Use It
        st.markdown("""
        ### 🔬 Understanding Lung Structure Analysis
        
        This tool automates the quantification of **alveolar structure** in lung histology images, providing 
        two critical morphometric measurements used in pulmonary research:
        
        - **MLI (Mean Linear Intercept)**: Estimates the average diameter of airspaces (alveoli).
        
        - **MWT (Mean Wall Thickness)**: Quantifies the thickness of alveolar septa (walls).
        
        #### 🎯 Why Use This Tool?
        
        Manual stereological measurements are time-consuming, operator-dependent, and prone to bias. This 
        automated approach:
        - **Increases reproducibility** by applying consistent thresholding and measurement criteria
        - **Saves time** by processing entire image sets in minutes
        - **Reduces bias** through objective, algorithm-based tissue segmentation
        - **Enables high-throughput analysis** of multiple regions of interest (ROIs) per specimen
        
        ---
        
        ### 📋 Step-by-Step User Guide
        
        #### 1️⃣ **Configure Analysis Parameters** (Left Sidebar)
        
        Before uploading images, set these critical parameters:
        
        - **Line Spacing (pixels)**: Distance between grid lines.
          
        - **Pixel Width & Height (µm)**: Your microscope calibration values. These convert pixel measurements 
          to real-world micrometers. Check your imaging software or microscope specifications.
          
        - **Slide-ROI Separator**: Character separating slide ID from ROI number in your filenames 
          (e.g., "HE0061-A_0001.tif" uses "_" as separator).
        
        #### 2️⃣ **Upload Your Images** (Upload & Process Tab)
        
        - Click the file uploader and select your H&E (or similar) stained lung histology images
        - Supported formats: TIF, TIFF, PNG, JPG, JPEG
        - You can upload multiple images simultaneously for batch processing
        - Name your output directory (default: "mli_output")
        
        **File Naming Convention**: Use the format `SlideID[separator]ROI.extension`
        - Example: `HE0061-A_0001.tif`, `Slide23_ROI05.png`
        - The separator helps the tool organize results by slide and calculate per-slide averages
        
        #### 3️⃣ **Run the Analysis**
        
        - Click "🚀 Start Processing"
        - Watch the progress bar and preview each processed image
        - Processing time depends on image size and number (typically 10-30 seconds per image)
        
        #### 4️⃣ **Review Results** (Results Tab)
        
        After processing completes:
        - **Summary Metrics**: View overall average MLI and MWT across all specimens
        - **Data Table**: Examine individual ROI measurements and per-slide averages
        - **Quality Control**: Check image previews to ensure proper tissue segmentation
        - **Download Options**: 
          - CSV files for statistical analysis in R, Python, or GraphPad Prism
          - Excel file with formatted results for reports
          - Complete ZIP archive with all processed images and raw data
        
        #### 5️⃣ **Interpreting Your Results**
        
        - **MLI Interpretation**:
          - ↑ MLI = Enlarged airspaces (emphysema, alveolar simplification)
          - ↓ MLI = Smaller airspaces (inflammation, atelectasis, developmental immaturity)
          
        - **MWT Interpretation**:
          - ↑ MWT = Thickened walls (fibrosis, inflammation, edema)
          - ↓ MWT = Thinned walls (emphysema, tissue degradation)
          
        - **Statistical Analysis**: Export the CSV and perform appropriate tests:
          - Compare groups with t-tests or ANOVA
          - Account for multiple ROIs per animal (nested/hierarchical designs)
          - Report both individual ROI values and per-specimen averages
        
        ---
        
        ### 🔧 How the Analysis Works (Technical Details)
        
        Understanding the computational workflow helps you interpret results and troubleshoot issues.
        
        #### Step 1: Image Thresholding (Tissue vs. Airspace Segmentation)
        
        The tool uses **Huang's Fuzzy Thresholding**, an entropy-based method that automatically determines 
        the optimal intensity value separating tissue (darker pixels) from airspaces (lighter pixels).
        
        **Why this method?**
        - Adapts to varying staining intensity across images
        - More robust than simple global thresholding
        - Handles gradual transitions at tissue boundaries (fuzzy logic)
        - Minimizes the ambiguity in pixel classification
        
        **What happens:**
        1. The algorithm analyzes the histogram of pixel intensities
        2. Calculates fuzzy membership functions for "tissue" and "air" classes
        3. Finds the threshold that minimizes classification entropy (uncertainty)
        4. Creates a binary image: white = airspace, black = tissue
        
        **Output**: `contrast_threshold.tif` - Review this to verify proper segmentation
        
        #### Step 2: Grid Overlay for Stereological Sampling
        
        Following classical stereology principles (Weibel 1963; American Thoracic Society guidelines), 
        the tool overlays two perpendicular grids:
        
        - **Horizontal grid**: Lines running left-to-right
        - **Vertical grid**: Lines running top-to-bottom
        
        **Why two orientations?**
        - Accounts for potential anisotropy (directional bias) in tissue structure
        - Provides more accurate average by sampling in all directions
        - Standard stereological practice to minimize geometric bias
        
        **Spacing**: Determined by your "Line Spacing" parameter.
        
        #### Step 3: Chord Length Measurement
        
        For **MLI** (airspace measurement):
        - Grid lines are overlaid on the INVERTED binary image (black = air, white = tissue)
        - The tool measures each continuous black segment where a grid line crosses airspace
        - These segments are "chords" - linear intercepts through airspace
        
        For **MWT** (wall thickness):
        - Grid lines are overlaid on the ORIGINAL binary image (white = air, black = tissue)
        - Measures continuous black segments representing tissue walls
        
        **Key processing steps:**
        1. Identify all contours (connected components) where grid lines intersect structures
        2. Measure the length of each contour using bounding box dimensions
        3. Filter out artifacts:
           - Exclude structures touching image edges (incomplete measurements)
           - Classify as horizontal or vertical based on aspect ratio
        4. Convert pixel lengths to micrometers using calibration values
        
        **Outputs per image:**
        - `horizontal_grid_mli.tif`, `vertical_grid_mli.tif`: Extracted MLI chords
        - `horizontal_grid_mwt.tif`, `vertical_grid_mwt.tif`: Extracted MWT chords
        - `overlay_horizontal_mli.tif`, etc.: Visual overlays showing measured segments in red
        - `chord_lengths_mli.csv`, `chord_lengths_mwt.csv`: Raw measurements with coordinates
        
        #### Step 4: Contour Analysis (Area Quantification)
        
        Beyond linear measurements, the tool quantifies total airspace area:
        
        1. **Contour Detection**: Identifies all enclosed white regions (airspaces) in the thresholded image
        2. **Edge Filtering**: Excludes contours touching image borders (incomplete alveoli)
        3. **Area Calculation**: Computes area in pixels, then converts to µm²
        4. **Visualization**: Generates images with red outlines showing detected contours
        
        **Outputs:**
        - `all_contours_[filename].tif`: Shows all detected contours
        - `final_contours_[filename].tif`: Shows filtered contours used for analysis
        - `image_summaries.csv`: Contains total airspace area per ROI
        
        **Use case**: Total airspace area helps assess overall tissue preservation and can be used to 
        normalize MLI/MWT values or calculate tissue fraction.
        
        #### Step 5: Statistical Aggregation
        
        The tool performs hierarchical data aggregation:
        
        1. **Per-Orientation Averaging**: 
           - Calculates mean chord length for horizontal chords
           - Calculates mean chord length for vertical chords
           - Separates MLI and MWT measurements
        
        2. **Per-ROI Averaging**:
           - Averages horizontal and vertical means to get final MLI per ROI
           - Averages horizontal and vertical means to get final MWT per ROI
           - Formula: `MLI_ROI = (mean_horizontal + mean_vertical) / 2`
        
        3. **Per-Slide Averaging**:
           - Groups all ROIs from the same slide (based on filename parsing)
           - Calculates average MLI and MWT across ROIs
           - Also averages total airspace area
        
        **Statistical outputs:**
        - `slide_set_grid_avg.csv`: Complete dataset with all levels of aggregation
        - `avg_per_slide.xlsx`: Formatted summary for each slide/specimen
        
        #### Algorithm Validation & Considerations
        
        **Strengths:**
        - Based on established stereological methods (Weibel, Knudsen et al.)
        - Automatic thresholding reduces operator bias
        - Batch processing ensures consistent methodology
        - Outputs include visual QC images for verification
        
        **Limitations & Best Practices:**
        - **Staining quality matters**: Poorly stained or over-stained sections may threshold incorrectly
        - **Image quality**: Ensure good focus and even illumination across the field
        - **Tissue artifacts**: Folding, tearing, or processing artifacts will affect measurements
        - **Grid density**: Too few lines = poor sampling; too many = overlapping measurements
        - **Edge effects**: ROIs should be well within tissue boundaries to avoid edge artifacts
        
        **Quality Control Recommendations:**
        1. Always review thresholded images - tissue should be black, airspaces white
        2. Check overlay images - red lines should trace tissue or airspace accurately
        3. Verify contour images - outlines should follow alveolar boundaries
        4. Compare automated results with manual measurements on a subset (validation)
        5. Look for outliers in the results table (processing failures)
        
        ---
        
        ### 📊 Output File Organization
        
        The tool creates a well-organized folder structure for each processed image to help you easily navigate and understand the results.
        
        **Organized Folder Structure** (in `output_dir/input_dir_name/image_name/`):
        
        ```
        image_name/
        ├── README.md                          # Documentation of folder structure
        ├── 01_preprocessing/
        │   └── contrast_threshold.tif         # Binary thresholded image
        ├── 02_contours/
        │   ├── all_contours_[name].tif        # All detected contours
        │   ├── final_contours_[name].tif      # Filtered contours (outlined)
        │   └── final_contours_filled_[name].tif  # Filtered contours (filled)
        ├── 03_mli/
        │   ├── grids/
        │   │   ├── horizontal_grid_mli.tif    # Horizontal MLI measurements
        │   │   └── vertical_grid_mli.tif      # Vertical MLI measurements
        │   └── overlays/
        │       ├── overlay_horizontal_mli.tif # MLI horizontal overlay
        │       └── overlay_vertical_mli.tif   # MLI vertical overlay
        ├── 04_mwt/
        │   ├── grids/
        │   │   ├── horizontal_grid_mwt.tif    # Horizontal MWT measurements
        │   │   └── vertical_grid_mwt.tif      # Vertical MWT measurements
        │   └── overlays/
        │       ├── overlay_horizontal_mwt.tif # MWT horizontal overlay
        │       └── overlay_vertical_mwt.tif   # MWT vertical overlay
        └── 05_data/
            ├── chord_lengths_mli.csv          # MLI raw measurements
            └── chord_lengths_mwt.csv          # MWT raw measurements
        ```
        
        **What Each Folder Contains:**
        
        - **01_preprocessing/**: The initial image processing step (thresholding to separate tissue from airspace)
        - **02_contours/**: Visualization of detected alveolar spaces
          - Outlined version shows boundaries
          - Filled version (red) shows which spaces were measured
        - **03_mli/**: Mean Linear Intercept analysis files
          - `grids/`: Raw grid line intersections with airspaces
          - `overlays/`: Visual QC - grid lines overlaid on tissue
        - **04_mwt/**: Mean Wall Thickness analysis files
          - `grids/`: Raw grid line intersections with tissue walls
          - `overlays/`: Visual QC - grid lines overlaid on tissue
        - **05_data/**: CSV files with individual chord measurements for statistical analysis
        
        **Summary Outputs** (in `output_dir/input_dir_name/`):
        - `image_summaries.csv`: Total areas and metadata per image
        
        **Aggregated Outputs** (in `output_dir/`):
        - `slide_set_grid_avg.csv`: Complete dataset with all measurements
        - `avg_per_slide.xlsx`: Per-specimen averages (ready for statistics)
        
        """)

        st.markdown("---")


if __name__ == "__main__":
    main()
